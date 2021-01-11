import argparse
import json
import logging.config
import os
import re

import openpyxl
import xlrd

_LOG_CONFIG = {
    'version': 1,
    'formatters': {
        'normal': {
            'format': '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        },
        'detailed': {
            'format': '%(asctime)s - %(name)s - %(module)s - %(levelname)s - %(message)s',
            'datefmt': '%Y-%m-%d %H:%M:%S'
        }
    },
    'handlers': {
        'console': {
            'class': 'logging.StreamHandler',
            'formatter': 'normal',
            'stream': 'ext://sys.stdout'
        },
        'file': {
            'class': 'logging.handlers.RotatingFileHandler',
            'formatter': 'detailed',
            'filename': './execution.log',
            'maxBytes': 1024,
            'backupCount': 3
        }
    },
    'loggers': {
        'excel_tools': {
            'level': 'DEBUG',
            'handlers': ['console']
        }
    }
}
logging.config.dictConfig(_LOG_CONFIG)
_logger = logging.getLogger('excel_tools')


class ExcelLoader:
    def __init__(self, path: str):
        self.path = path
        self.name = ''
        self.sheets = list()
        self.loaded = False
        if os.path.exists(path) and os.path.isfile(path) and re.match(r'.*\.xlsx?$', path):
            self.name = os.path.basename(path)
            if self.path.endswith('.xls'):
                self._load_xls()
            if self.path.endswith('.xlsx'):
                self._load_xlsx()
        if self.sheets:
            self.loaded = True
        if not self.loaded:
            _logger.error('Load {} failed.'.format(self.path))

    def _load_xls(self):
        """
        Load Excel with .xls type to a Python list. This program loads all sheets by default. Each sheet is converted
        into a Python dictionary. The List is formatted like following:
        [{'name': 'sheet1', 'headers': ['h1', 'h2'], 'contents':[{'h1': cell(2,1), 'h2': cell(2,2)}]}, {sheet2}...]
        Note: The first row of each sheet is treated as the headers of each sheet. Hence, the first row must have the
        same formatting with the rest of rows.
        :return:
        """
        wb = xlrd.open_workbook(self.path)
        _logger.info('Loading {} (this Excel file contains {} sheets)...'.format(self.path, wb.nsheets))
        for sheet_idx in range(wb.nsheets):
            current_sheet = wb.sheet_by_index(sheet_idx)
            current_sheet_dict = dict()
            current_sheet_dict['name'] = current_sheet.name
            current_sheet_dict['headers'] = tuple(current_sheet.row_values(0, 0))
            current_sheet_contents = list()
            for row_idx in range(1, current_sheet.nrows):
                current_sheet_contents.append(dict(zip(current_sheet_dict['headers'],
                                                       current_sheet.row_values(row_idx, 0))))
            current_sheet_dict['contents'] = current_sheet_contents
            self.sheets.append(current_sheet_dict)
        _logger.info('{} successfully loaded, {} sheets in total.'.format(self.path, wb.nsheets))

    def _load_xlsx(self):
        """
        Load Excel with .xlsx type to a Python list. This program loads all sheets by default. Each sheet is converted
        into a Python dictionary. The List is formatted like following:
        [{'name': 'sheet1', 'headers': ['h1', 'h2'], 'contents':[{'h1': cell(2,1), 'h2': cell(2,2)}]}, {sheet2}...]
        Note: The first row of each sheet is treated as the headers of each sheet. Hence, the first row must have the
        same formatting with the rest of rows.
        :return:
        """
        wb = openpyxl.load_workbook(self.path)
        _logger.info('Loading {} (this Excel file contains {} sheets)...'.format(self.path, len(wb.sheetnames)))
        for sheet in wb:
            current_sheet_dict = dict()
            current_sheet_dict['name'] = sheet.title
            current_sheet_dict['headers'] = list(sheet.iter_rows(min_row=1, max_row=1, min_col=1,
                                                                 max_col=sheet.max_column, values_only=True))[0]
            current_sheet_contents = list()
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1,
                                       max_col=sheet.max_column, values_only=True):
                current_sheet_contents.append(dict(zip(current_sheet_dict['headers'], row)))
            current_sheet_dict['contents'] = current_sheet_contents
            self.sheets.append(current_sheet_dict)
        _logger.info('{} successfully loaded, {} sheets in total.'.format(self.path, len(wb.sheetnames)))

    def dumps(self):
        """
        Convert ExcelClass into Json, and save the Json into a file in the same path with input Excel.
        :return:
        """
        if self.loaded:
            json_dict = dict()
            json_dict['name'] = self.name
            json_dict['path'] = self.path
            json_dict['sheets'] = self.sheets
            json_path = os.path.join(os.path.dirname(self.path), self.name[:self.name.rfind('.')] + '.json')
            with open(file=json_path, mode='w', encoding='utf-8') as f:
                f.write(json.dumps(json_dict, ensure_ascii=False))
                f.flush()
                f.close()
            _logger.info('{} successfully loaded, dumped to {}'.format(self.path, json_path))
        else:
            _logger.error('Load {} failed, can not be dumped into Json.'.format(self.path))


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Frequently used tools for Excel files with .xlsx type support, '
                                                 'e.g., converting excel to Json, combining multiple Excel files.',
                                     epilog='Enjoy this program! :)',
                                     allow_abbrev=False)
    parser.add_argument('action', type=str, choices=['2json', 'combine'],
                        help='available functions')
    parser.add_argument('-if', '--input_file', type=str, dest='input_file',
                        help='path to the input Excel file')
    parser.add_argument('--dumps', action='store_true', dest='dumps',
                        help='dumps loaded Json to file in the same path with input Excel file')
    args = parser.parse_args()

    print('2json' == args.action)
    if '2json' == args.action and args.input_file:
        excel = ExcelLoader(args.input_file)
        if args.dumps:
            excel.dumps()
